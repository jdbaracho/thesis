"""Persist a redactor translation table as an Excel workbook.

The translation table produced by :meth:`src.pdf_redactor.PDFRedactor.redact`
maps ``entity_text`` → ``{"id": "PERSON-1", "scores": {"PERSON": 0.85, ...}}``.
This module renders that structure as a human-readable ``.xlsx`` with one
merged row per entity_text and one sub-row per (entity_type, score) pair.
"""

from __future__ import annotations

from pathlib import Path
from typing import Union

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


__all__ = ["save_translation_table_xlsx"]


def save_translation_table_xlsx(
    translation_table: dict,
    output_xlsx: Union[str, Path],
) -> None:
    """Persist ``translation_table`` as an Excel workbook at ``output_xlsx``.

    Each ``entity_text`` occupies one (merged) row, with its
    ``entity_type``/``score`` pairs laid out as sub-rows beside it.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "translation_table"

    headers = ["entity_text", "id", "entity_type", "score"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    center = Alignment(horizontal="center", vertical="center")
    row_idx = 2
    for entity_text, entry in translation_table.items():
        scores = entry["scores"]
        if not scores:
            continue
        start_row = row_idx
        for entity_type, score in scores.items():
            ws.cell(row=row_idx, column=1, value=entity_text)
            ws.cell(row=row_idx, column=2, value=entry["id"])
            ws.cell(row=row_idx, column=3, value=entity_type)
            ws.cell(row=row_idx, column=4, value=score)
            row_idx += 1
        end_row = row_idx - 1
        if end_row > start_row:
            ws.merge_cells(
                start_row=start_row, start_column=1,
                end_row=end_row, end_column=1,
            )
            ws.merge_cells(
                start_row=start_row, start_column=2,
                end_row=end_row, end_column=2,
            )
        ws.cell(row=start_row, column=1).alignment = center
        ws.cell(row=start_row, column=2).alignment = center

    wb.save(str(output_xlsx))
