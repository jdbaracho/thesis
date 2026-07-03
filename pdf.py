from typing import List

# For Presidio
from presidio_analyzer import AnalyzerEngine, RecognizerResult
from custom_extensions.custom_image_analyzer import CustomImageAnalyzerEngine

# For PDF redaction
import fitz  # PyMuPDF

# For image conversion
import io
from PIL import Image, ImageDraw, ImageFont

# For pretty printing
import pprint

# For Excel export
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from custom_extensions.custom_utils import resolve_conflicts

pp = pprint.PrettyPrinter()

# LLM Recognizer
from presidio_extensions.basic_langextract_recognizer import BasicLangExtractRecognizer

analyzer = AnalyzerEngine()
analyzer.registry.add_recognizer(BasicLangExtractRecognizer(config_path="presidio_extensions/config/ollama_config.yaml"))
image_analyzer = CustomImageAnalyzerEngine(analyzer_engine=analyzer)
# llm_recognizer = BasicLangExtractRecognizer(config_path="extensions/config/ollama_config.yaml")

# translation_table[entity_text] = {
#     "id": <str | None>,            # e.g. "PERSON-1"; filled in by finalize_translation_table()
#     "scores": {entity_type: score} # all candidate types detected for this entity_text
# }
translation_table = {}


def process_results(results: List[RecognizerResult], text: str) -> None:
    for result in results:
        entity_text = text[result.start:result.end]
        if entity_text not in translation_table:
            translation_table[entity_text] = {
                "id": None,
                "scores": {result.entity_type: result.score},
            }
        else:
            scores = translation_table[entity_text]["scores"]
            if result.entity_type not in scores:
                scores[result.entity_type] = result.score
            else:
                scores[result.entity_type] = max(scores[result.entity_type], result.score)


def finalize_translation_table() -> None:
    """Assign an `id` to every entry: the highest-scoring entity_type for
    that entity_text, suffixed with a sequential counter per entity_type
    (in insertion order across the translation_table)."""
    type_counters: dict = {}
    for entity_text, entry in translation_table.items():
        scores = entry["scores"]
        if not scores:
            entry["id"] = None
            continue
        top_type = max(scores.items(), key=lambda kv: kv[1])[0]
        type_counters[top_type] = type_counters.get(top_type, 0) + 1
        entry["id"] = f"{top_type}-{type_counters[top_type]}"

doc = fitz.open("input/text_image.pdf")

# Collected during analysis, applied after finalize_translation_table()
# so we can label each box with its assigned id (e.g. "PERSON-1").
# Each entry: (page, fitz.Rect, entity_text)
pending_redactions: list = []

# Image redactions are also deferred so that, after finalize_translation_table()
# runs, every box can be drawn with its assigned id label (matching the text
# section's behaviour). Each entry: (page, xref, pil_image, [(box, entity_text), ...])
pending_image_redactions: list = []

for page in doc:
    # --- Redact images ---
    for img_info in page.get_images(full=True):
        xref = img_info[0]
        img_data = doc.extract_image(xref)

        pil_image = Image.open(io.BytesIO(img_data["image"]))

        bboxes, text = image_analyzer.analyze(
          pil_image,
        )
        process_results(bboxes, text)

        pp.pprint(bboxes)

        if not bboxes:
            continue

        # Capture entity_text for each box now (we have the OCR'd `text`);
        # actual drawing/replacing happens after ids are assigned.
        image_entries = [
            (box, text[box.start:box.end]) for box in bboxes
        ]
        pending_image_redactions.append((page, xref, pil_image, image_entries))

    # --- Redact text ---
    text_dict = page.get_text("rawdict")

    for block in text_dict["blocks"]:
        if block["type"] != 0:  # Skip non-text blocks (e.g. images)
            continue

        for line in block["lines"]:
            for span in line["spans"]:
                chars = span.get("chars", [])
                text = "".join(c["c"] for c in chars)

                if not text.strip() or not chars:
                    continue

                # Analyze the span text with Presidio
                results = analyzer.analyze(text=text, language='en')
                results = resolve_conflicts(text, results)

                # results = llm_recognizer.analyze(text=text)

                process_results(results, text)

                # decision_process = results[0].analysis_explanation
                # pp.pprint(decision_process.__dict__)

                for result in results:
                    matched_chars = chars[result.start:result.end]
                    if not matched_chars:
                        continue

                    # Combine individual character bounding boxes
                    x0 = min(c["bbox"][0] for c in matched_chars)
                    y0 = min(c["bbox"][1] for c in matched_chars)
                    x1 = max(c["bbox"][2] for c in matched_chars)
                    y1 = max(c["bbox"][3] for c in matched_chars)

                    entity_text = text[result.start:result.end]
                    pending_redactions.append(
                        (page, fitz.Rect(x0, y0, x1, y1), entity_text)
                    )

# Compute ids now that all results have been collected.
finalize_translation_table()

# --- Draw labeled redaction boxes on images and push them back into the PDF.
# Matches the text section: white fill + black centered label sized to the box.
def _load_font(size: int) -> ImageFont.ImageFont:
    for candidate in ("Helvetica.ttc", "Arial.ttf", "DejaVuSans.ttf"):
        try:
            return ImageFont.truetype(candidate, size)
        except (IOError, OSError):
            continue
    return ImageFont.load_default()

for page, xref, pil_image, image_entries in pending_image_redactions:
    draw = ImageDraw.Draw(pil_image)
    for box, entity_text in image_entries:
        x0 = box.left
        y0 = box.top
        x1 = x0 + box.width
        y1 = y0 + box.height
        # change here and remove draw.text if you want to just black out the text without any label
        draw.rectangle([x0, y0, x1, y1], fill=(255, 255, 255))

        label = translation_table.get(entity_text, {}).get("id") or ""
        if not label:
            continue

        fontsize = max(8, int(box.height * 0.6))
        font = _load_font(fontsize)
        tx0, ty0, tx1, ty1 = font.getbbox(label)
        tw, th = tx1 - tx0, ty1 - ty0
        tx = x0 + (box.width - tw) / 2 - tx0
        ty = y0 + (box.height - th) / 2 - ty0
        draw.text((tx, ty), label, fill=(0, 0, 0), font=font)

    output = io.BytesIO()
    pil_image.save(output, format="PNG")
    page.replace_image(xref, stream=output.getvalue())

# Add redaction annotations with their labels, then apply them per page.
pages_to_apply = set()
for page, rect, entity_text in pending_redactions:
    label = translation_table.get(entity_text, {}).get("id") or ""
    # Scale font to fit the box height; PyMuPDF will clip if the label is
    # wider than the rect (entity ids are often longer than the original text).
    fontsize = max(1.0, rect.height * 0.6)
    page.add_redact_annot(
        rect,
        text=label,
        fontname="helv",
        fontsize=fontsize,
        text_color=(0, 0, 0),
        align=fitz.TEXT_ALIGN_CENTER,
    )

    # use this instead of the above if you want to just black out the text without any label
    # page.add_redact_annot(fitz.Rect(x0, y0, x1, y1), fill=(0, 0, 0))

    pages_to_apply.add(page)

for page in pages_to_apply:
    # Apply redactions: removes the underlying text from the content stream
    # and draws the labeled black rectangles in its place.
    page.apply_redactions()

output_pdf = "output/text_image_redacted.pdf"
os.makedirs(os.path.dirname(output_pdf), exist_ok=True)
doc.save(output_pdf)

# Save translation table as Excel alongside the redacted PDF.
# Each entity_text occupies one (merged) row, with its entity_type/score pairs
# laid out as sub-rows beside it.
output_xlsx = os.path.splitext(output_pdf)[0] + ".xlsx"
wb = Workbook()
ws = wb.active
ws.title = "translation_table"

headers = ["entity_text", "id", "entity_type", "score"]
ws.append(headers)
for cell in ws[1]:
    cell.font = Font(bold=True)

center = Alignment(horizontal="center", vertical="center")
row_idx = 2
for entity_text, entry in translation_table.items():
    scores = entry["scores"]
    if not scores:
        continue
    start_row = row_idx
    for entity_type, score in scores.items():
        ws.cell(row=row_idx, column=1, value=entity_text)
        ws.cell(row=row_idx, column=2, value=entry["id"])
        ws.cell(row=row_idx, column=3, value=entity_type)
        ws.cell(row=row_idx, column=4, value=score)
        row_idx += 1
    end_row = row_idx - 1
    if end_row > start_row:
        ws.merge_cells(start_row=start_row, start_column=1,
                       end_row=end_row, end_column=1)
        ws.merge_cells(start_row=start_row, start_column=2,
                       end_row=end_row, end_column=2)
    ws.cell(row=start_row, column=1).alignment = center
    ws.cell(row=start_row, column=2).alignment = center

wb.save(output_xlsx)
