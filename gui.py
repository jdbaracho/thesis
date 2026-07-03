"""Tkinter GUI for batch PDF redaction using Presidio + PyMuPDF."""

import io
import os
import sys
import threading
import traceback
from pathlib import Path
from typing import List
from tkinter import Tk, BooleanVar, Listbox, StringVar, filedialog, messagebox, scrolledtext
from tkinter import ttk


def _resource_root() -> Path:
    """Return the directory containing bundled resources (works in dev + PyInstaller)."""
    if getattr(sys, "frozen", False):
        # PyInstaller onedir: resources live next to the executable in Contents/Resources
        meipass = getattr(sys, "_MEIPASS", None)
        if meipass:
            return Path(meipass)
    return Path(__file__).resolve().parent


def _configure_tesseract() -> None:
    """Point pytesseract at the bundled tesseract binary, if present."""
    root = _resource_root()
    bundled_bin = root / "tesseract_bin" / "tesseract"
    bundled_tessdata = root / "tesseract_bin" / "tessdata"
    if bundled_bin.exists():
        try:
            import pytesseract  # noqa: WPS433
            pytesseract.pytesseract.tesseract_cmd = str(bundled_bin)
        except Exception:
            pass
        if bundled_tessdata.exists():
            os.environ["TESSDATA_PREFIX"] = str(bundled_tessdata)


_configure_tesseract()

import fitz  # PyMuPDF  # noqa: E402
from PIL import Image  # noqa: E402
from presidio_analyzer import AnalyzerEngine, RecognizerResult  # noqa: E402
from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment, Font  # noqa: E402

from custom_extensions.custom_image_analyzer import CustomImageRedactorEngine  # noqa: E402


SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_OUTPUT = Path.home() / "Downloads"


def _process_results(
    results: List[RecognizerResult],
    text: str,
    translation_table: dict,
) -> None:
    """Accumulate detected entities (text -> {entity_type: max_score})."""
    for result in results:
        entity_text = text[result.start:result.end]
        if entity_text not in translation_table:
            translation_table[entity_text] = {result.entity_type: result.score}
        else:
            entry = translation_table[entity_text]
            if result.entity_type not in entry:
                entry[result.entity_type] = result.score
            else:
                entry[result.entity_type] = max(entry[result.entity_type], result.score)


def _save_translation_table(translation_table: dict, output_xlsx: Path) -> None:
    """Persist the translation table as an Excel workbook."""
    wb = Workbook()
    ws = wb.active
    ws.title = "translation_table"

    headers = ["entity_text", "entity_type", "score"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    center = Alignment(horizontal="center", vertical="center")
    row_idx = 2
    for entity_text, types in translation_table.items():
        if not types:
            continue
        start_row = row_idx
        for entity_type, score in types.items():
            ws.cell(row=row_idx, column=1, value=entity_text)
            ws.cell(row=row_idx, column=2, value=entity_type)
            ws.cell(row=row_idx, column=3, value=score)
            row_idx += 1
        end_row = row_idx - 1
        if end_row > start_row:
            ws.merge_cells(
                start_row=start_row, start_column=1,
                end_row=end_row, end_column=1,
            )
        ws.cell(row=start_row, column=1).alignment = center

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_xlsx))


def redact_pdf(input_path: Path, output_path: Path, analyzer: AnalyzerEngine,
               image_redactor: CustomImageRedactorEngine, log) -> dict:
    """Redact a single PDF file (text + images) and return its translation table."""
    doc = fitz.open(str(input_path))
    translation_table: dict = {}

    for page_idx, page in enumerate(doc, start=1):
        # --- Redact images ---
        for img_info in page.get_images(full=True):
            xref = img_info[0]
            try:
                img_data = doc.extract_image(xref)
                pil_image = Image.open(io.BytesIO(img_data["image"]))
                redacted_pil, image_results, image_text = image_redactor.redact(
                    pil_image, (0, 0, 0)
                )
                _process_results(image_results, image_text, translation_table)
                buf = io.BytesIO()
                redacted_pil.save(buf, format="PNG")
                page.replace_image(xref, stream=buf.getvalue())
            except Exception as exc:  # noqa: BLE001
                log(f"  ! image xref={xref} on page {page_idx} skipped: {exc}")

        # --- Redact text ---
        text_dict = page.get_text("rawdict")
        for block in text_dict["blocks"]:
            if block.get("type") != 0:
                continue
            for line in block["lines"]:
                for span in line["spans"]:
                    chars = span.get("chars", [])
                    text = "".join(c["c"] for c in chars)
                    if not text.strip() or not chars:
                        continue

                    results = analyzer.analyze(text=text, language="en")
                    _process_results(results, text, translation_table)
                    for result in results:
                        matched_chars = chars[result.start:result.end]
                        if not matched_chars:
                            continue
                        x0 = min(c["bbox"][0] for c in matched_chars)
                        y0 = min(c["bbox"][1] for c in matched_chars)
                        x1 = max(c["bbox"][2] for c in matched_chars)
                        y1 = max(c["bbox"][3] for c in matched_chars)
                        page.add_redact_annot(
                            fitz.Rect(x0, y0, x1, y1), fill=(0, 0, 0)
                        )

        page.apply_redactions()

    output_path.parent.mkdir(parents=True, exist_ok=True)
    doc.save(str(output_path))
    doc.close()
    return translation_table


class RedactorApp:
    def __init__(self, root: Tk) -> None:
        self.root = root
        root.title("PDF Redactor")
        root.geometry("720x520")

        self.selected_files: list[Path] = []
        self.output_dir = StringVar(value=str(DEFAULT_OUTPUT))
        self.generate_xlsx = BooleanVar(value=True)
        self.status = StringVar(value="Ready.")
        self._worker: threading.Thread | None = None

        self._build_ui()

    def _build_ui(self) -> None:
        pad = {"padx": 8, "pady": 4}
        frm = ttk.Frame(self.root)
        frm.pack(fill="both", expand=True, padx=10, pady=10)

        # Input files
        ttk.Label(frm, text="Input PDFs:").grid(row=0, column=0, sticky="nw", **pad)
        files_frame = ttk.Frame(frm)
        files_frame.grid(row=0, column=1, sticky="nsew", **pad)
        self.files_list = Listbox(files_frame, height=5, activestyle="none")
        files_scroll = ttk.Scrollbar(
            files_frame, orient="vertical", command=self.files_list.yview
        )
        self.files_list.configure(yscrollcommand=files_scroll.set)
        self.files_list.pack(side="left", fill="both", expand=True)
        files_scroll.pack(side="right", fill="y")

        files_btns = ttk.Frame(frm)
        files_btns.grid(row=0, column=2, sticky="nw", **pad)
        ttk.Button(files_btns, text="Choose files…", command=self._pick_input).pack(
            fill="x"
        )
        ttk.Button(files_btns, text="Choose folder…", command=self._pick_input_folder).pack(
            fill="x", pady=(4, 0)
        )
        ttk.Button(files_btns, text="Clear", command=self._clear_files).pack(
            fill="x", pady=(4, 0)
        )

        # Output folder
        ttk.Label(frm, text="Output folder:").grid(row=1, column=0, sticky="w", **pad)
        ttk.Entry(frm, textvariable=self.output_dir).grid(
            row=1, column=1, sticky="ew", **pad
        )
        ttk.Button(frm, text="Browse…", command=self._pick_output).grid(
            row=1, column=2, **pad
        )

        # Options
        ttk.Checkbutton(
            frm,
            text="Generate Excel translation table (.xlsx)",
            variable=self.generate_xlsx,
        ).grid(row=2, column=1, sticky="w", **pad)

        # Run button
        self.run_btn = ttk.Button(frm, text="Redact PDFs", command=self._start)
        self.run_btn.grid(row=3, column=0, columnspan=3, pady=(10, 4), sticky="ew")

        # Progress
        self.progress = ttk.Progressbar(frm, mode="determinate")
        self.progress.grid(row=4, column=0, columnspan=3, sticky="ew", **pad)

        # Status
        ttk.Label(frm, textvariable=self.status, foreground="#555").grid(
            row=5, column=0, columnspan=3, sticky="w", **pad
        )

        # Log
        ttk.Label(frm, text="Log:").grid(row=6, column=0, sticky="w", **pad)
        ttk.Button(frm, text="Clear log", command=self._clear_log).grid(
            row=6, column=2, sticky="e", **pad
        )
        self.log_widget = scrolledtext.ScrolledText(frm, height=15, state="disabled")
        self.log_widget.grid(row=7, column=0, columnspan=3, sticky="nsew", **pad)

        frm.columnconfigure(1, weight=1)
        frm.rowconfigure(7, weight=1)

    # ------ UI helpers ------
    def _pick_input(self) -> None:
        paths = filedialog.askopenfilenames(
            title="Select PDF files",
            initialdir=str(Path.home()),
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
        )
        if paths:
            self.selected_files = [Path(p) for p in paths]
            self._refresh_files_list()

    def _pick_input_folder(self) -> None:
        folder = filedialog.askdirectory(
            title="Select folder with PDFs",
            initialdir=str(Path.home()),
        )
        if folder:
            pdfs = sorted(p for p in Path(folder).glob("*.pdf") if p.is_file())
            if not pdfs:
                messagebox.showinfo("No PDFs", f"No .pdf files found in:\n{folder}")
                return
            self.selected_files = pdfs
            self._refresh_files_list()

    def _clear_files(self) -> None:
        self.selected_files = []
        self._refresh_files_list()

    def _refresh_files_list(self) -> None:
        self.files_list.delete(0, "end")
        for p in self.selected_files:
            self.files_list.insert("end", p.name)

    def _pick_output(self) -> None:
        path = filedialog.askdirectory(initialdir=self.output_dir.get() or str(SCRIPT_DIR))
        if path:
            self.output_dir.set(path)

    def _log(self, msg: str) -> None:
        def append() -> None:
            self.log_widget.configure(state="normal")
            self.log_widget.insert("end", msg + "\n")
            self.log_widget.see("end")
            self.log_widget.configure(state="disabled")

        self.root.after(0, append)

    def _clear_log(self) -> None:
        self.log_widget.configure(state="normal")
        self.log_widget.delete("1.0", "end")
        self.log_widget.configure(state="disabled")

    def _set_status(self, msg: str) -> None:
        self.root.after(0, lambda: self.status.set(msg))

    # ------ Worker ------
    def _start(self) -> None:
        if self._worker and self._worker.is_alive():
            return

        out_dir = Path(self.output_dir.get()).expanduser()

        pdfs = [p for p in self.selected_files if p.is_file() and p.suffix.lower() == ".pdf"]
        if not pdfs:
            messagebox.showinfo("No PDFs", "Please choose one or more PDF files.")
            return

        try:
            out_dir.mkdir(parents=True, exist_ok=True)
        except OSError as exc:
            messagebox.showerror("Output error", f"Cannot create output folder:\n{exc}")
            return

        self.run_btn.configure(state="disabled")
        self.progress.configure(maximum=len(pdfs), value=0)
        self._worker = threading.Thread(
            target=self._run, args=(pdfs, out_dir), daemon=True
        )
        self._worker.start()

    def _run(self, pdfs: list[Path], out_dir: Path) -> None:
        try:
            self._set_status("Loading Presidio engines…")
            self._log("Initializing analyzer and image redactor…")
            analyzer = AnalyzerEngine()
            image_redactor = CustomImageRedactorEngine()

            for i, pdf in enumerate(pdfs, start=1):
                self._set_status(f"Processing {i}/{len(pdfs)}: {pdf.name}")
                self._log(f"[{i}/{len(pdfs)}] {pdf.name}")
                out_path = out_dir / f"{pdf.stem}_redacted.pdf"
                try:
                    translation_table = redact_pdf(
                        pdf, out_path, analyzer, image_redactor, self._log
                    )
                    self._log(f"  ✓ saved {out_path.name}")
                    if self.generate_xlsx.get():
                        xlsx_path = out_path.with_suffix(".xlsx")
                        try:
                            _save_translation_table(translation_table, xlsx_path)
                            self._log(f"  ✓ saved {xlsx_path.name}")
                        except Exception as exc:  # noqa: BLE001
                            self._log(f"  ! translation table not saved: {exc}")
                except Exception as exc:  # noqa: BLE001
                    self._log(f"  ✗ failed: {exc}")
                    self._log(traceback.format_exc())
                self.root.after(0, lambda v=i: self.progress.configure(value=v))

            self._set_status(f"Done. {len(pdfs)} file(s) processed.")
            self._log("Finished.")
            self.root.after(
                0,
                lambda: messagebox.showinfo(
                    "Done", f"Processed {len(pdfs)} file(s).\nOutput: {out_dir}"
                ),
            )
        except Exception as exc:  # noqa: BLE001
            self._log(traceback.format_exc())
            self._set_status("Error.")
            self.root.after(0, lambda: messagebox.showerror("Error", str(exc)))
        finally:
            self.root.after(0, lambda: self.run_btn.configure(state="normal"))


def main() -> None:
    os.chdir(SCRIPT_DIR)
    root = Tk()
    RedactorApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
